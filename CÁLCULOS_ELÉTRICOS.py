import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl import Workbook
import math
import pathlib
from io import BytesIO
import os  # Adicionado para manipula√ß√£o de caminhos
import plotly.express as px

st.set_page_config(layout="wide", page_title="Dimensionamento El√©trico")

# Define o caminho do arquivo na mesma pasta do script
CAMINHO_ARQUIVO = os.path.join(os.path.dirname(__file__), "Quadro_de_cargas.xlsx")

# Fun√ß√µes do backend adaptadas para Streamlit
def criar_planilha_se_nao_existir():
    minha_borda = Side(border_style="thin", color='000000')
    
    if not pathlib.Path(CAMINHO_ARQUIVO).exists():
        ficheiro = Workbook()
        folha = ficheiro.active
        folha.title = "QD"
    else:
        ficheiro = openpyxl.load_workbook(CAMINHO_ARQUIVO)
        if "QD" not in ficheiro.sheetnames:
            folha = ficheiro.create_sheet("QD")
        else:
            return  # Se j√° existe, n√£o precisa fazer nada

    folha = ficheiro["QD"]
    
    headers = [
        ("A1", "N¬∞"), ("B1", "DESCRI√á√ÉO"), ("C1", "ATIVA-R"), ("D1", "ATIVA-S"), ("E1", "ATIVA-T"),
        ("F1", "DEM-R"), ("G1", "DEM-S"), ("H1", "DEM-T"), ("I1", "R"), ("J1", "S"), ("K1", "T"), ("L1", "FP"),
        ("M1", "FD"), ("N1", "TENS√ÉO FASE (V)"), ("O1", "TENS√ÉO LINHA (V)"), ("P1", "POT. TOTAL (W)"),
        ("Q1", "DEM. TOTAL (VA)"), ("R1", "COR. M√âDIA (A)"), ("S1", "DIST.(M)"), ("T1", "QUEDA DE TENS√ÉO (%)"),
        ("U1", "FA"), ("V1", "NE"), ("W1", "TE"), ("X1", "DISJUNTOR")
    ]
    
    for cell, value in headers:
        folha[cell] = value
        folha[cell].font = Font(color='ffffff', bold=True, size=12)
        folha[cell].fill = PatternFill('solid', start_color="162B4E")
        folha[cell].border = Border(top=minha_borda, left=minha_borda, right=minha_borda, bottom=minha_borda)

    ficheiro.save(CAMINHO_ARQUIVO)

def calcular_dimensionamento(nome_quadro, fp, fd, dist, pr, ps, pt, tensao):
    cb_voltenax_095_qd = [7.54, 4.5, 2.86, 1.83, 1.34, 1, 0.71, 0.53, 0.43, 0.36]
    cb_voltenax_bitola = [6, 10, 16, 25, 35, 50, 70, 95, 120, 150]
    cb_voltenax_terra = [6, 10, 16, 16, 16, 25, 35, 50, 70, 95]
    cb_voltenax_corrente = [54, 75, 100, 133, 164, 198, 253, 306, 354, 407]
    dj_cx_mol = [40, 50, 63, 100, 125, 150, 160, 200, 250, 320, 400, 500, 630, 700, 800, 1000, 1600, 2000, 2500]
    
    sum_pot = pr + ps + pt
    p = [pr, ps, pt]
    c_qds = []
    
    # C√°lculo trif√°sico
    if pr != 0 and ps != 0 and pt != 0:
        c_med = sum_pot / (tensao * (math.sqrt(3)) * fp)
        for i in range(3):
            c_qds.append(c_med * (p[i] / (sum_pot / 3)))
    
    # C√°lculo bif√°sico
    elif pr == 0 or ps == 0 or pt == 0:
        c_med = sum_pot / (tensao * fp)
        for i in range(3):
            c_qds.append(c_med * (p[i] / (sum_pot / 2)))
    
    # C√°lculo monof√°sico
    else:
        c_med = sum_pot / ((tensao * fp) / math.sqrt(3))
        for i in range(3):
            c_qds.append(c_med * (p[i] / sum_pot))
    
    # Determinar queda de tens√£o
    qd = [(dist * r * c_med) / (10 * tensao) for r in cb_voltenax_095_qd]

    n = 0
    cabo = ""
    terra = ""
    queda = 0
    dj = 0  # Inicializa o disjuntor
    max_cabos = 5  # Altere conforme o n√∫mero m√°ximo de cabos em paralelo que deseja testar

    # Tenta de 1 at√© max_cabos cabos
    for n_cabos in range(1, max_cabos + 1):
        for i in range(len(cb_voltenax_corrente)):
            corrente_limite = cb_voltenax_corrente[i] * n_cabos
            queda_total = qd[i] / n_cabos
            if c_med < corrente_limite and queda_total < 3:  # ou 3.5 se quiser afrouxar
                queda = queda_total
                cabo = f"{n_cabos}x{cb_voltenax_bitola[i]}"
                terra = cb_voltenax_terra[i]  # pode mudar se quiser 1 terra por fase
                n = n_cabos
                break
        if n > 0:
            break                                                                                               

    # Disjuntor
    if c_med < 32:
        dj = 32
    else:
        for i in range(len(dj_cx_mol)):
            if c_med < dj_cx_mol[i]:
                dj = dj_cx_mol[i]
                break
    
    # Tens√£o de linha
    tensao_linha = 127 if tensao == 220 else 220
    
    # Gera o n√∫mero do QD
    qd_number = 1
    try:
        wb = openpyxl.load_workbook(CAMINHO_ARQUIVO)
        if 'QD' in wb.sheetnames:
            qd_number = wb['QD'].max_row
    except:
        pass

    # Retorna os dados na ordem CORRETA correspondente √†s colunas do Excel
    return {
        "N¬∞": f"QD-{qd_number}",
        "DESCRI√á√ÉO": nome_quadro,
        "ATIVA-R": pr,
        "ATIVA-S": ps,
        "ATIVA-T": pt,
        "DEM-R": fd * pr,
        "DEM-S": fd * ps,
        "DEM-T": fd * pt,
        "R": c_qds[0],
        "S": c_qds[1],
        "T": c_qds[2],
        "FP": fp,
        "FD": fd,
        "TENS√ÉO FASE (V)": tensao,
        "TENS√ÉO LINHA (V)": tensao_linha,
        "POT. TOTAL (W)": sum_pot,
        "DEM. TOTAL (VA)": (sum_pot*fd) / fp,
        "COR. M√âDIA (A)": c_med,
        "DIST.(M)": dist,
        "QUEDA DE TENS√ÉO (%)": queda,
        "FA": cabo,        # Fase (condutor fase)
        "NE": cabo,        # Neutro (mesmo condutor que a fase)
        "TE": terra,       # Terra (condutor de prote√ß√£o)
        "DISJUNTOR": dj   # Disjuntor
    }

def salvar_no_excel(dados):
    # Verifica se o arquivo existe, se n√£o, cria
    if not pathlib.Path(CAMINHO_ARQUIVO).exists():
        criar_planilha_se_nao_existir()
    
    ficheiro = openpyxl.load_workbook(CAMINHO_ARQUIVO)
    
    # Verifica se a planilha QD existe, se n√£o, cria
    if 'QD' not in ficheiro.sheetnames:
        criar_planilha_se_nao_existir()
        ficheiro = openpyxl.load_workbook(CAMINHO_ARQUIVO)
    
    folha = ficheiro["QD"]
    
    nova_linha = folha.max_row + 1
    
    for col, valor in enumerate(dados.values(), start=1):
        folha.cell(row=nova_linha, column=col, value=valor)
    
    ficheiro.save(CAMINHO_ARQUIVO)
    st.success("Dados salvos com sucesso!")

def carregar_dados():
    try:
        df = pd.read_excel(CAMINHO_ARQUIVO, 
                          sheet_name="QD",
                          header=0,
                          engine='openpyxl')
        
        if len(df) > 0:
            return df
        return pd.DataFrame()
    except Exception as e:
        st.warning(f"Erro ao carregar dados: {str(e)}")
        return pd.DataFrame()
    
# Interface do Streamlit
st.title("Dimensionamento de Quadros de Carga")

# Criar abas
tab1, tab2, tab3 = st.tabs(["C√°lculo", "Visualiza√ß√£o", "Dados Salvos"])

with tab1:
    st.header("C√°lculo de Quadro de Cargas")
    
    with st.form("form_calculo"):
        col1, col2 = st.columns(2)
        
        with col1:
            nome_quadro = st.text_input("Nome do Quadro:")
            dist = st.number_input("Dist√¢ncia do QGBT at√© o Quadro (m):", min_value=0.1, step=0.1)
            fp = st.selectbox("Fator de Pot√™ncia:", [0.92, 0.8, 0.75, 0.7], index=0)
            fd = st.selectbox("Fator de Demanda:", [1.0, 0.9, 0.8, 0.7, 0.6, 0.5], index=0)
        
        with col2:
            tensao = st.selectbox("Tens√£o (V):", [220, 380], index=0)
            pr = st.number_input("Pot√™ncia - R (W):", min_value=0.0, step=100.0)
            ps = st.number_input("Pot√™ncia - S (W):", min_value=0.0, step=100.0)
            pt = st.number_input("Pot√™ncia - T (W):", min_value=0.0, step=100.0)
        
        if st.form_submit_button("Calcular e Salvar"):
            # Primeiro cria a planilha se n√£o existir
            criar_planilha_se_nao_existir()
    
            # Depois calcula os dados
            dados = calcular_dimensionamento(nome_quadro, fp, fd, dist, pr, ps, pt, tensao)
    
            # Finalmente salva
            salvar_no_excel(dados)

with tab2:
    st.header("Visualiza√ß√£o de Resultados")
    
    df = carregar_dados()
    if not df.empty:
        # Mostrar dataframe completo
        st.dataframe(df)
        
        # Criar abas para os gr√°ficos
        tab_graficos1, tab_graficos2 = st.tabs(["An√°lise de Pot√™ncia e Demanda", "An√°lise de Corrente e Queda de Tens√£o"])
        
        with tab_graficos1:

            st.subheader("Pot√™ncia vs Demanda por Fase")

             # Preparar os dados somados
            potencias = df[["ATIVA-R", "ATIVA-S", "ATIVA-T"]].sum()
            demandas = df[["DEM-R", "DEM-S", "DEM-T"]].sum()

             # Montar DataFrame no formato longo (long format)
            fases = ["Fase R", "Fase S", "Fase T"]
            df_plot = pd.DataFrame({
                "Fase": fases * 2,
                "Valor (kW ou kVA)": list(potencias / 1000) + list(demandas / 1000),
                "Tipo": ["Pot√™ncia"] * 3 + ["Demanda"] * 3
            })

            fig = px.bar(
                df_plot,
                x="Fase",
                y="Valor (kW ou kVA)",
                color="Tipo",
                barmode="group",
                text_auto=".2s",
                title="Compara√ß√£o entre Pot√™ncia Instalada e Demanda por Fase"
            )
            fig.update_layout(
                xaxis_title="Fase",
                yaxis_title="Valor (kW ou kVA)",
                legend_title="Tipo",
                bargap=0.2
            )
            st.plotly_chart(fig, use_container_width=True)

            st.caption("Gr√°fico de barras agrupadas comparando as pot√™ncias ativas e demandas por fase.")

        
        with tab_graficos2:
            col1, col2 = st.columns(2)
            
            with col1:
                # Gr√°fico de Corrente por Fase
                st.subheader("Corrente por Fase")
                
                # Preparar dados de corrente
                dados_corrente = df[["R", "S", "T"]].sum()
                dados_corrente.index = ["Fase R", "Fase S", "Fase T"]
                
                # Plotar gr√°fico
                st.bar_chart(dados_corrente.rename("Corrente (A)"))
                
                # Adicionar estat√≠sticas
                st.metric("Corrente M√©dia Total", f"{dados_corrente.mean():.2f} A")
            
            with col2:
                # Gr√°fico de Queda de Tens√£o
                st.subheader("Queda de Tens√£o por Circuito")
                
                # Preparar dados
                dados_queda = df.set_index("DESCRI√á√ÉO")["QUEDA DE TENS√ÉO (%)"]
                
                # Plotar gr√°fico
                st.bar_chart(dados_queda)
                
                # Adicionar estat√≠sticas
                st.metric("Queda de Tens√£o M√°xima", f"{dados_queda.max():.2f} %")
        
        # Adicionar se√ß√£o de an√°lise geral
        st.subheader("An√°lise Geral")

        # Criar m√©tricas resumidas
        col_met1, col_met2, col_met3, col_met4 = st.columns(4)

        # Calcular totais
        potencia_total = df['POT. TOTAL (W)'].sum()
        demanda_total = df['DEM. TOTAL (VA)'].sum()
        corrente_media = df['COR. M√âDIA (A)'].mean()

        # Tamanhos padr√£o de subesta√ß√µes (em kVA)
        subestacoes = [75, 112.5, 225, 300, 500, 750, 1000, 1250, 1500, 1750, 2000]
        demanda_kva = demanda_total / 1000  # Converter VA para kVA

        # Encontrar a subesta√ß√£o adequada
        subestacao_recomendada = min([s for s in subestacoes if s >= demanda_kva], default=subestacoes[-1])

        with col_met1:
            st.metric("Pot√™ncia Total Instalada", f"{potencia_total:,.2f} W")

        with col_met2:
            st.metric("Demanda Total Calculada", f"{demanda_total:,.2f} VA")

        with col_met3:
            st.metric("Corrente M√©dia Total", f"{corrente_media:.2f} A")

        with col_met4:
            st.metric("Subesta√ß√£o Recomendada", 
                    f"{subestacao_recomendada} kVA",
                    help=f"Baseado na demanda total de {demanda_kva:.2f} kVA")

        # Adicionar gr√°fico de compara√ß√£o com as subesta√ß√µes
        st.subheader("Dimensionamento da Subesta√ß√£o")

        # Criar DataFrame para o gr√°fico
        df_subestacao = pd.DataFrame({
            'Capacidade (kVA)': subestacoes,
            'Tipo': 'Dispon√≠vel'
        })

        # Adicionar demanda atual
        df_demanda = pd.DataFrame({
            'Capacidade (kVA)': [demanda_kva],
            'Tipo': 'Demanda Calculada'
        })

        df_plot = pd.concat([df_subestacao, df_demanda])

        # Plotar gr√°fico
        fig = px.bar(df_plot, 
                    x='Capacidade (kVA)', 
                    y='Tipo', 
                    color='Tipo',
                    orientation='h',
                    title=f'Demanda Calculada: {demanda_kva:.2f} kVA vs Capacidades de Subesta√ß√£o',
                    text='Capacidade (kVA)',
                    height=400)

        # Destacar a recomendada
        fig.add_vline(x=subestacao_recomendada, line_width=2, line_dash="dash", line_color="green",
                    annotation_text=f"Recomendado: {subestacao_recomendada} kVA", 
                    annotation_position="top right")

        # Ajustar layout
        fig.update_layout(showlegend=False)
        fig.update_traces(texttemplate='%{text:.0f} kVA', textposition='outside')

        st.plotly_chart(fig, use_container_width=True)

        # Adicionar explica√ß√£o
        st.info(f"""
        **Legenda:**
        - üü¶ Barras azuis: Capacidades padr√£o de subesta√ß√µes
        - üüß Barra laranja: Sua demanda calculada ({demanda_kva:.2f} kVA)
        - üü© Linha verde: Subesta√ß√£o recomendada ({subestacao_recomendada} kVA)

        A subesta√ß√£o recomendada √© a menor capacidade padr√£o que atende ou excede sua demanda calculada.
        """)
    else:
        st.warning("Nenhum dado encontrado. Realize c√°lculos primeiro.")
with tab3:
    st.header("Dados Salvos no Excel")
    
    df = carregar_dados()  # J√° carrega corretamente
    
    if not df.empty:
        st.dataframe(df)
        
        # Op√ß√£o para baixar
        try:
            CAMINHO_ARQUIVO = os.path.join(os.path.dirname(__file__), "Quadro_de_cargas.xlsx")
            output = BytesIO()
            wb = openpyxl.load_workbook(CAMINHO_ARQUIVO)
            wb.save(output)
            output.seek(0)

            st.download_button(
                label="Baixar dados como Excel",
                data=output,
                file_name="quadros_de_carga.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Erro ao preparar download: {str(e)}")

        # Apagar arquivos caso necess√°rio
        st.subheader("Excluir Quadros de Carga")

        opcoes_quadro = df["DESCRI√á√ÉO"].tolist()
        quadro_selecionado = st.selectbox("Selecione o Quadro que deseja excluir:", opcoes_quadro)

        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("Apagar Quadro Selecionado"):
                # Remove do DataFrame
                df_filtrado = df[df["DESCRI√á√ÉO"] != quadro_selecionado]

                # Salva novamente o Excel (mantendo cabe√ßalho)
                try:
                    CAMINHO_ARQUIVO = os.path.join(os.path.dirname(__file__), "Quadro_de_cargas.xlsx")
                    with pd.ExcelWriter(CAMINHO_ARQUIVO, engine='openpyxl') as writer:
                        df_filtrado.to_excel(writer, sheet_name='QD', index=False)
                    st.success(f"Quadro '{quadro_selecionado}' apagado com sucesso.")
                    st.experimental_rerun()  # Atualiza a p√°gina para mostrar os dados atualizados
                except Exception as e:
                    st.error(f"Erro ao excluir: {str(e)}")
        
        with col2:
            if st.button("üóëÔ∏è Apagar TODOS os Quadros", type="primary"):
                try:
                    # Cria um novo arquivo Excel com apenas o cabe√ßalho
                    wb = Workbook()
                    folha = wb.active
                    folha.title = "QD"
                    
                    # Adiciona os cabe√ßalhos
                    headers = [
                        "N¬∞", "DESCRI√á√ÉO", "ATIVA-R", "ATIVA-S", "ATIVA-T",
                        "DEM-R", "DEM-S", "DEM-T", "R", "S", "T", "FP",
                        "FD", "TENS√ÉO FASE (V)", "TENS√ÉO LINHA (V)", "POT. TOTAL (W)",
                        "DEM. TOTAL (VA)", "COR. M√âDIA (A)", "DIST.(M)", "QUEDA DE TENS√ÉO (%)",
                        "FA", "NE", "TE", "DISJUNTOR"
                    ]
                    
                    folha.append(headers)
                    
                    # Formata√ß√£o do cabe√ßalho
                    minha_borda = Side(border_style="thin", color='000000')
                    for cell in folha[1]:
                        cell.font = Font(color='ffffff', bold=True, size=12)
                        cell.fill = PatternFill('solid', start_color="162B4E")
                        cell.border = Border(top=minha_borda, left=minha_borda, right=minha_borda, bottom=minha_borda)
                    
                    wb.save(CAMINHO_ARQUIVO)
                    st.success("Todos os quadros foram apagados com sucesso!")
                    st.experimental_rerun()  # Atualiza a p√°gina para mostrar os dados atualizados
                except Exception as e:
                    st.error(f"Erro ao apagar todos os quadros: {str(e)}")
        
        st.warning("Aten√ß√£o: Apagar todos os quadros √© uma a√ß√£o irrevers√≠vel. Todos os dados ser√£o perdidos, mantendo apenas o cabe√ßalho da planilha.")
    else:
        st.info("Nenhum dado salvo ainda. Realize c√°lculos na aba 'C√°lculo'.")
